import io
from datetime import datetime
from openpyxl import load_workbook
from ..extensions import db
from ..models.intersection import Intersection, TrafficLight, ElectronicPolice
from ..models.point import ParkingEnforcementPoint, CheckpointPoint, ParkingEnforcement, Checkpoint
from ..models.project import Project


class ExcelImportService:
    @staticmethod
    def import_excel(file_storage, import_type: str) -> dict:
        wb = load_workbook(file_storage)
        count = 0
        errors = []

        if import_type == 'intersection':
            count, errors = ExcelImportService._import_intersections(wb)
        elif import_type == 'parking_point':
            c1, e1 = ExcelImportService._import_parking_enforcement_points(wb)
            count, errors = c1, e1
        elif import_type == 'checkpoint_point':
            c1, e1 = ExcelImportService._import_checkpoint_points(wb)
            count, errors = c1, e1
        elif import_type == 'project':
            count, errors = ExcelImportService._import_projects(wb)
        elif import_type == 'device':
            count, errors = ExcelImportService._import_devices(wb)
        else:
            raise ValueError(f'不支持的导入类型: {import_type}')

        return {'count': count, 'errors': errors}

    @staticmethod
    def _import_projects(wb) -> tuple:
        count = 0
        errors = []

        if '项目' not in wb.sheetnames:
            errors.append('未找到"项目"工作表')
            return count, errors

        ws = wb['项目']
        headers = [cell.value for cell in ws[1]]

        project_name_idx = headers.index('项目名称') if '项目名称' in headers else -1
        contract_amount_idx = headers.index('合同金额') if '合同金额' in headers else -1
        acceptance_date_idx = headers.index('验收日期') if '验收日期' in headers else -1
        warranty_period_idx = headers.index('质保期') if '质保期' in headers else -1
        warranty_expire_date_idx = headers.index('质保到期时间') if '质保到期时间' in headers else -1
        builder_idx = headers.index('建设单位') if '建设单位' in headers else -1
        constructor_idx = headers.index('施工单位') if '施工单位' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                project_name = row[project_name_idx] if project_name_idx >= 0 else None
                if not project_name:
                    continue

                existing = db.session.query(Project).filter_by(name=project_name).first()
                if existing:
                    continue

                project = Project()
                project.name = project_name
                if contract_amount_idx >= 0 and row[contract_amount_idx]:
                    project.contract_amount = row[contract_amount_idx]
                if acceptance_date_idx >= 0 and row[acceptance_date_idx]:
                    project.acceptance_date = row[acceptance_date_idx]
                if warranty_period_idx >= 0 and row[warranty_period_idx]:
                    project.warranty_period = str(row[warranty_period_idx])
                if warranty_expire_date_idx >= 0 and row[warranty_expire_date_idx]:
                    if isinstance(row[warranty_expire_date_idx], datetime):
                        project.warranty_expire_date = row[warranty_expire_date_idx].date()
                    else:
                        project.warranty_expire_date = row[warranty_expire_date_idx]
                if builder_idx >= 0:
                    project.builder = row[builder_idx]
                if constructor_idx >= 0:
                    project.construction_unit = row[constructor_idx]

                db.session.add(project)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_intersections(wb) -> tuple:
        count = 0
        errors = []

        if '路口' not in wb.sheetnames:
            errors.append('未找到"路口"工作表')
            return count, errors

        ws = wb['路口']
        headers = [cell.value for cell in ws[1]]

        name_idx = headers.index('路口名称') if '路口名称' in headers else -1
        type_idx = headers.index('路口类型') if '路口类型' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                intersection_name = row[name_idx] if name_idx >= 0 else None
                if not intersection_name:
                    continue

                existing = db.session.query(Intersection).filter_by(name=intersection_name).first()
                if existing:
                    continue

                intersection = Intersection()
                intersection.name = intersection_name
                if type_idx >= 0:
                    intersection.type = row[type_idx]

                db.session.add(intersection)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_parking_enforcement_points(wb) -> tuple:
        count = 0
        errors = []

        sheet_name = '违停点位' if '违停点位' in wb.sheetnames else '点位'
        if sheet_name not in wb.sheetnames:
            errors.append('未找到"违停点位"工作表')
            return count, errors

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        name_idx = headers.index('点位名称') if '点位名称' in headers else -1
        area_idx = headers.index('抓拍区域') if '抓拍区域' in headers else -1
        type_idx = headers.index('安装位置') if '安装位置' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                point_name = row[name_idx] if name_idx >= 0 else None
                if not point_name:
                    continue

                existing = db.session.query(ParkingEnforcementPoint).filter_by(name=point_name).first()
                if existing:
                    continue

                point = ParkingEnforcementPoint()
                point.name = point_name
                if area_idx >= 0:
                    point.area = row[area_idx]
                if type_idx >= 0:
                    point.type = row[type_idx]

                db.session.add(point)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_checkpoint_points(wb) -> tuple:
        count = 0
        errors = []

        sheet_name = '卡口点位' if '卡口点位' in wb.sheetnames else '点位'
        if sheet_name not in wb.sheetnames:
            return count, errors

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        name_idx = headers.index('点位名称') if '点位名称' in headers else -1
        area_idx = headers.index('卡口类型') if '卡口类型' in headers else -1
        type_idx = headers.index('安装位置') if '安装位置' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                point_name = row[name_idx] if name_idx >= 0 else None
                if not point_name:
                    continue

                existing = db.session.query(CheckpointPoint).filter_by(name=point_name).first()
                if existing:
                    continue

                point = CheckpointPoint()
                point.name = point_name
                if area_idx >= 0:
                    point.area = row[area_idx]
                if type_idx >= 0:
                    point.type = row[type_idx]

                db.session.add(point)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_devices(wb) -> tuple:
        count = 0
        errors = []

        traffic_light_count = 0
        ep_count = 0
        pe_count = 0
        cp_count = 0

        if '信号灯' in wb.sheetnames:
            c, e = ExcelImportService._import_traffic_lights(wb['信号灯'])
            traffic_light_count = c
            errors.extend(e)

        if '电子警察' in wb.sheetnames:
            c, e = ExcelImportService._import_electronic_police(wb['电子警察'])
            ep_count = c
            errors.extend(e)

        if '违停球' in wb.sheetnames:
            c, e = ExcelImportService._import_parking_enforcement(wb['违停球'])
            pe_count = c
            errors.extend(e)

        if '卡口' in wb.sheetnames:
            c, e = ExcelImportService._import_checkpoint(wb['卡口'])
            cp_count = c
            errors.extend(e)

        count = traffic_light_count + ep_count + pe_count + cp_count
        return count, errors

    @staticmethod
    def _import_traffic_lights(ws) -> tuple:
        count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]

        intersection_name_idx = headers.index('路口名称') if '路口名称' in headers else -1
        project_name_idx = headers.index('归属项目') if '归属项目' in headers else -1
        signal_type_idx = headers.index('信号机类型') if '信号机类型' in headers else -1
        signal_count_idx = headers.index('信号机数量') if '信号机数量' in headers else -1
        left_arrow_idx = headers.index('左转箭头灯数量') if '左转箭头灯数量' in headers else -1
        straight_arrow_idx = headers.index('直行箭头数量') if '直行箭头数量' in headers else -1
        right_arrow_idx = headers.index('右转箭头数量') if '右转箭头数量' in headers else -1
        full_screen_idx = headers.index('满屏灯数量') if '满屏灯数量' in headers else -1
        non_motor_idx = headers.index('非机动灯数量') if '非机动灯数量' in headers else -1
        pedestrian_idx = headers.index('人行灯数量') if '人行灯数量' in headers else -1
        radar_idx = headers.index('车流量雷达数量') if '车流量雷达数量' in headers else -1
        guide_screen_idx = headers.index('诱导屏数量') if '诱导屏数量' in headers else -1
        power_source_idx = headers.index('取电说明') if '取电说明' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                intersection_name = row[intersection_name_idx] if intersection_name_idx >= 0 else None
                project_name = row[project_name_idx] if project_name_idx >= 0 else None

                if not intersection_name or not project_name:
                    continue

                intersection = db.session.query(Intersection).filter_by(name=intersection_name).first()
                project = db.session.query(Project).filter_by(name=project_name).first()

                if not intersection or not project:
                    errors.append(f'第{row_idx}行: 未找到路口或项目')
                    continue

                tl = TrafficLight()
                tl.intersection_id = intersection.id
                tl.project_id = project.id

                if signal_type_idx >= 0:
                    tl.signal_type = row[signal_type_idx]
                if signal_count_idx >= 0 and row[signal_count_idx]:
                    tl.signal_count = int(row[signal_count_idx])
                if left_arrow_idx >= 0 and row[left_arrow_idx]:
                    tl.left_arrow_count = int(row[left_arrow_idx])
                if straight_arrow_idx >= 0 and row[straight_arrow_idx]:
                    tl.straight_arrow_count = int(row[straight_arrow_idx])
                if right_arrow_idx >= 0 and row[right_arrow_idx]:
                    tl.right_arrow_count = int(row[right_arrow_idx])
                if full_screen_idx >= 0 and row[full_screen_idx]:
                    tl.full_screen_count = int(row[full_screen_idx])
                if non_motor_idx >= 0 and row[non_motor_idx]:
                    tl.non_motor_count = int(row[non_motor_idx])
                if pedestrian_idx >= 0 and row[pedestrian_idx]:
                    tl.pedestrian_count = int(row[pedestrian_idx])
                if radar_idx >= 0 and row[radar_idx]:
                    tl.radar_count = int(row[radar_idx])
                if guide_screen_idx >= 0 and row[guide_screen_idx]:
                    tl.guide_screen_count = int(row[guide_screen_idx])
                if power_source_idx >= 0:
                    tl.power_source = row[power_source_idx]

                db.session.add(tl)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行信号灯导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_electronic_police(ws) -> tuple:
        count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]

        intersection_name_idx = headers.index('路口名称') if '路口名称' in headers else -1
        project_name_idx = headers.index('归属项目') if '归属项目' in headers else -1
        capture_type_idx = headers.index('抓拍类型') if '抓拍类型' in headers else -1
        terminal_server_idx = headers.index('终端服务器数量') if '终端服务器数量' in headers else -1
        forward_idx = headers.index('正向抓拍数量') if '正向抓拍数量' in headers else -1
        reverse_idx = headers.index('反向抓拍数量') if '反向抓拍数量' in headers else -1
        led_idx = headers.index('LED灯') if 'LED灯' in headers else -1
        strobe_idx = headers.index('爆闪灯') if '爆闪灯' in headers else -1
        ptz_idx = headers.index('监控球机数量') if '监控球机数量' in headers else -1
        detector_idx = headers.index('信号检测器数量') if '信号检测器数量' in headers else -1
        network_idx = headers.index('取网说明') if '取网说明' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                intersection_name = row[intersection_name_idx] if intersection_name_idx >= 0 else None
                project_name = row[project_name_idx] if project_name_idx >= 0 else None

                if not intersection_name or not project_name:
                    continue

                intersection = db.session.query(Intersection).filter_by(name=intersection_name).first()
                project = db.session.query(Project).filter_by(name=project_name).first()

                if not intersection or not project:
                    errors.append(f'第{row_idx}行: 未找到路口或项目')
                    continue

                ep = ElectronicPolice()
                ep.intersection_id = intersection.id
                ep.project_id = project.id

                if capture_type_idx >= 0:
                    ep.capture_type = row[capture_type_idx]
                if terminal_server_idx >= 0 and row[terminal_server_idx]:
                    ep.terminal_server_count = int(row[terminal_server_idx])
                if forward_idx >= 0 and row[forward_idx]:
                    ep.forward_capture_count = int(row[forward_idx])
                if reverse_idx >= 0 and row[reverse_idx]:
                    ep.reverse_capture_count = int(row[reverse_idx])
                if led_idx >= 0 and row[led_idx]:
                    ep.led_light_count = int(row[led_idx])
                if strobe_idx >= 0 and row[strobe_idx]:
                    ep.strobe_light_count = int(row[strobe_idx])
                if ptz_idx >= 0 and row[ptz_idx]:
                    ep.ptz_count = int(row[ptz_idx])
                if detector_idx >= 0 and row[detector_idx]:
                    ep.signal_detector_count = int(row[detector_idx])
                if network_idx >= 0:
                    ep.network_source = row[network_idx]

                db.session.add(ep)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行电子警察导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_parking_enforcement(ws) -> tuple:
        count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]

        point_name_idx = headers.index('点位名称') if '点位名称' in headers else -1
        project_name_idx = headers.index('归属项目') if '归属项目' in headers else -1
        camera_idx = headers.index('抓拍机数量') if '抓拍机数量' in headers else -1
        sign_idx = headers.index('违停标牌数量') if '违停标牌数量' in headers else -1
        monitor_sign_idx = headers.index('监控标牌数量') if '监控标牌数量' in headers else -1
        power_idx = headers.index('取电说明') if '取电说明' in headers else -1
        network_idx = headers.index('取网说明') if '取网说明' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                point_name = row[point_name_idx] if point_name_idx >= 0 else None
                project_name = row[project_name_idx] if project_name_idx >= 0 else None

                if not point_name or not project_name:
                    continue

                point = db.session.query(ParkingEnforcementPoint).filter_by(name=point_name).first()
                project = db.session.query(Project).filter_by(name=project_name).first()

                if not point or not project:
                    errors.append(f'第{row_idx}行: 未找到违停点位或项目')
                    continue

                pe = ParkingEnforcement()
                pe.point_id = point.id
                pe.project_id = project.id

                if camera_idx >= 0 and row[camera_idx]:
                    pe.camera_count = int(row[camera_idx])
                if sign_idx >= 0 and row[sign_idx]:
                    pe.parking_sign_count = int(row[sign_idx])
                if monitor_sign_idx >= 0 and row[monitor_sign_idx]:
                    pe.monitor_sign_count = int(row[monitor_sign_idx])
                if power_idx >= 0:
                    pe.power_source = row[power_idx]
                if network_idx >= 0:
                    pe.network_source = row[network_idx]

                db.session.add(pe)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行违停抓拍导入失败: {str(e)}')

        return count, errors

    @staticmethod
    def _import_checkpoint(ws) -> tuple:
        count = 0
        errors = []
        headers = [cell.value for cell in ws[1]]

        point_name_idx = headers.index('点位名称') if '点位名称' in headers else -1
        project_name_idx = headers.index('归属项目') if '归属项目' in headers else -1
        camera_idx = headers.index('抓拍机数量') if '抓拍机数量' in headers else -1
        strobe_idx = headers.index('爆闪灯数量') if '爆闪灯数量' in headers else -1
        radar_idx = headers.index('测速雷达数量') if '测速雷达数量' in headers else -1
        sign_idx = headers.index('标牌数量') if '标牌数量' in headers else -1
        power_idx = headers.index('取电说明') if '取电说明' in headers else -1
        network_idx = headers.index('取网说明') if '取网说明' in headers else -1

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue

            try:
                point_name = row[point_name_idx] if point_name_idx >= 0 else None
                project_name = row[project_name_idx] if project_name_idx >= 0 else None

                if not point_name or not project_name:
                    continue

                point = db.session.query(CheckpointPoint).filter_by(name=point_name).first()
                project = db.session.query(Project).filter_by(name=project_name).first()

                if not point or not project:
                    errors.append(f'第{row_idx}行: 未找到卡口点位或项目')
                    continue

                cp = Checkpoint()
                cp.point_id = point.id
                cp.project_id = project.id

                if camera_idx >= 0 and row[camera_idx]:
                    cp.camera_count = int(row[camera_idx])
                if strobe_idx >= 0 and row[strobe_idx]:
                    cp.strobe_light_count = int(row[strobe_idx])
                if radar_idx >= 0 and row[radar_idx]:
                    cp.radar_count = int(row[radar_idx])
                if sign_idx >= 0 and row[sign_idx]:
                    cp.sign_count = int(row[sign_idx])
                if power_idx >= 0:
                    cp.power_source = row[power_idx]
                if network_idx >= 0:
                    cp.network_source = row[network_idx]

                db.session.add(cp)
                db.session.commit()
                count += 1
            except Exception as e:
                db.session.rollback()
                errors.append(f'第{row_idx}行卡口导入失败: {str(e)}')

        return count, errors
