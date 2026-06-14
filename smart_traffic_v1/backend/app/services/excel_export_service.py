import io
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from flask import send_file
from ..extensions import db
from ..models.intersection import Intersection, TrafficLight, ElectronicPolice
from ..models.point import ParkingEnforcementPoint, CheckpointPoint, ParkingEnforcement, Checkpoint, SkyNetPoint, SkyNet
from ..models.project import Project
from ..models.backend_device import BackendDevice


class ExcelExportService:
    @staticmethod
    def _get_project_info(project_id):
        project = db.session.query(Project).get(project_id)
        if not project:
            return {
                'name': '',
                'acceptance_date': '',
                'warranty_period': '',
                'warranty_expire_date': '',
                'builder': '',
                'construction_unit': ''
            }
        return {
            'name': project.name or '',
            'acceptance_date': project.acceptance_date.isoformat() if project.acceptance_date else '',
            'warranty_period': project.warranty_period or '',
            'warranty_expire_date': project.warranty_expire_date.isoformat() if project.warranty_expire_date else '',
            'builder': project.builder or '',
            'construction_unit': project.construction_unit or ''
        }

    @staticmethod
    def _get_warranty_expire_date(project_id):
        project = db.session.query(Project).get(project_id)
        return project.warranty_expire_date if project else None

    @staticmethod
    def _warranty_closer_to_now(date1, date2):
        """判断 date2 是否比 date1 更靠近今天"""
        if date1 is None: return True
        if date2 is None: return False
        today = date.today()
        diff1 = abs((date1 - today).days)
        diff2 = abs((date2 - today).days)
        return diff2 < diff1

    @staticmethod
    def _should_keep_new_item(existing_item, new_item):
        """去重判断：保留质保到期日期更靠近当前的记录（通用）"""
        existing_date = ExcelExportService._get_warranty_expire_date(existing_item.project_id)
        new_date = ExcelExportService._get_warranty_expire_date(new_item.project_id)
        return ExcelExportService._warranty_closer_to_now(existing_date, new_date)

    @staticmethod
    def _should_keep_for_service(existing_item, new_item, selected_map=None):
        """服役期限去重：优先保留选择了项目的记录"""
        # 优先保留选中状态（与前端 ServiceRanking 的 is_selected 对齐）
        if selected_map is not None:
            existing_selected = selected_map.get(existing_item.id, False)
            new_selected = selected_map.get(new_item.id, False)
            if new_selected and not existing_selected:
                return True
            if existing_selected and not new_selected:
                return False
        # 其次优先保留有 project_id 的记录
        if new_item.project_id and not existing_item.project_id:
            return True
        if not new_item.project_id and existing_item.project_id:
            return False
        return ExcelExportService._should_keep_new_item(existing_item, new_item)

    @staticmethod
    def _get_warranty_status(project_id):
        project = db.session.query(Project).get(project_id)
        if not project or not project.warranty_expire_date:
            return '点位无关联项目'
        if project.warranty_expire_date >= date.today():
            return '在保'
        return '过保'

    @staticmethod
    def _get_intersection_type_name(intersection_id):
        intersection = db.session.query(Intersection).get(intersection_id)
        return intersection.type if intersection else ''

    @staticmethod
    def _get_intersection_name(intersection_id):
        intersection = db.session.query(Intersection).get(intersection_id)
        return intersection.name if intersection else ''

    @staticmethod
    def _get_point_info(point_id, table='parking_enforcement'):
        model_cls = ParkingEnforcementPoint if table == 'parking_enforcement' else CheckpointPoint
        point = db.session.query(model_cls).get(point_id)
        if not point:
            return {'name': '', 'area': '', 'type': ''}
        return {
            'name': point.name or '',
            'area': point.area or '',
            'type': point.type or ''
        }

    @staticmethod
    def _get_backend_device_type_name(device_id):
        device = db.session.query(BackendDevice).get(device_id)
        return device.type if device else ''

    @staticmethod
    def export_statistics() -> send_file:
        wb = Workbook()
        wb.remove(wb.active)

        ExcelExportService._create_project_overview_sheet(wb)
        ExcelExportService._create_traffic_light_sheet(wb)
        ExcelExportService._create_electronic_police_sheet(wb)
        ExcelExportService._create_parking_enforcement_sheet(wb)
        ExcelExportService._create_checkpoint_sheet(wb)
        ExcelExportService._create_sky_net_sheet(wb)
        ExcelExportService._create_backend_device_sheet(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'智能交通设备统计_{date.today()}.xlsx'
        )

    @staticmethod
    def download_template() -> send_file:
        wb = Workbook()
        wb.remove(wb.active)

        ExcelExportService._create_traffic_light_template(wb)
        ExcelExportService._create_electronic_police_template(wb)
        ExcelExportService._create_parking_enforcement_template(wb)
        ExcelExportService._create_checkpoint_template(wb)
        ExcelExportService._create_backend_device_template(wb)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='智能交通数据导入模板.xlsx'
        )

    @staticmethod
    def _create_traffic_light_template(wb):
        ws = wb.create_sheet('信号灯')
        headers = [
            '路口名称', '路口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '信号机类型', '信号机数量', '左转箭头灯数量', '直行箭头数量',
            '右转箭头数量', '满屏灯数量', '非机动灯数量', '人行灯数量',
            '车流量雷达数量', '诱导屏数量', '取电说明'
        ]
        ws.append(headers)
        ws.append(['', '', '', '', '', '', '', '', '', '', 0, 0, 0, 0, 0, 0, 0, 0, 0, ''])
        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_electronic_police_template(wb):
        ws = wb.create_sheet('电子警察')
        headers = [
            '路口名称', '路口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍类型', '终端服务器数量', '正向抓拍数量', '反向抓拍数量',
            'LED灯', '爆闪灯', '监控球机数量', '信号检测器数量', '取网说明'
        ]
        ws.append(headers)
        ws.append(['', '', '', '', '', '', '', '', '', '', 0, 0, 0, 0, 0, 0, 0, ''])
        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_parking_enforcement_template(wb):
        ws = wb.create_sheet('违停球')
        headers = [
            '点位名称', '抓拍区域', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍机数量', '违停标牌数量', '监控标牌数量', '取电说明', '取网说明'
        ]
        ws.append(headers)
        ws.append(['', '', '', '', '', '', '', '', '', 0, 0, 0, '', ''])
        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_checkpoint_template(wb):
        ws = wb.create_sheet('卡口')
        headers = [
            '点位名称', '卡口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍机数量', '爆闪灯数量', '测速雷达数量', '标牌数量', '取电说明', '取网说明'
        ]
        ws.append(headers)
        ws.append(['', '', '', '', '', '', '', '', '', 0, 0, 0, 0, '', ''])
        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_backend_device_template(wb):
        ws = wb.create_sheet('后端设备')
        headers = [
            '设备名称', '品牌型号', '设备类型', '设备数量', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位'
        ]
        ws.append(headers)
        ws.append(['', '', '', '', '', '', '', '', '', '', ''])
        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_project_overview_sheet(wb):
        ws = wb.create_sheet('项目概览')
        headers = [
            '序号', '项目名称', '建设单位', '信号灯', '电子警察', '违停球', 
            '卡口', '结构化相机', '后端设备', '合计', '质保到期', '质保状态'
        ]
        ws.append(headers)

        projects = db.session.query(Project).all()
        
        # 先收集所有项目数据
        rows = []
        for project in projects:
            tl_count = db.session.query(TrafficLight).filter_by(project_id=project.id).count()
            ep_count = db.session.query(ElectronicPolice).filter_by(project_id=project.id).count()
            pe_count = db.session.query(ParkingEnforcement).filter_by(project_id=project.id).count()
            cp_count = db.session.query(Checkpoint).filter_by(project_id=project.id).count()
            sn_count = db.session.query(SkyNet).filter_by(project_id=project.id).count()
            bd_count = db.session.query(BackendDevice).filter_by(project_id=project.id).count()
            
            total = tl_count + ep_count + pe_count + cp_count + sn_count + bd_count
            
            warranty_expire_date = project.warranty_expire_date.isoformat() if project.warranty_expire_date else '-'
            warranty_status = '在保' if project.warranty_expire_date and project.warranty_expire_date >= date.today() else '过保' if project.warranty_expire_date else '-'

            rows.append({
                'name': project.name or '',
                'builder': project.builder or '-',
                'tl': tl_count, 'ep': ep_count, 'pe': pe_count,
                'cp': cp_count, 'sn': sn_count, 'bd': bd_count,
                'total': total,
                'warranty_expire_date': warranty_expire_date,
                'warranty_status': warranty_status
            })
        
        # 排序：质保状态优先（过保 > 在保 > 其他），再按建设单位
        status_order = {'过保': 0, '在保': 1}
        rows.sort(key=lambda r: (
            status_order.get(r['warranty_status'], 2),
            r['builder']
        ))
        
        for idx, r in enumerate(rows, 1):
            ws.append([
                idx, r['name'], r['builder'],
                r['tl'], r['ep'], r['pe'], r['cp'], r['sn'], r['bd'],
                r['total'], r['warranty_expire_date'], r['warranty_status']
            ])

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_traffic_light_sheet(wb):
        ws = wb.create_sheet('信号灯')
        headers = [
            '序号', '路口名称', '路口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '信号机类型', '信号机数量', '左转箭头灯数量', '直行箭头数量',
            '右转箭头数量', '满屏灯数量', '非机动灯数量', '人行灯数量',
            '倒计时器数量', '车流量雷达数量', '诱导屏数量', '取电说明', '设备服役时长（年）'
        ]
        ws.append(headers)

        traffic_lights = db.session.query(TrafficLight).all()
        intersections = db.session.query(Intersection).all()
        
        # 构建选中状态映射
        tl_selected_map = {}
        for inter in intersections:
            if inter.selected_traffic_light_id:
                tl_selected_map[inter.selected_traffic_light_id] = True
        
        # 通用去重（质保日期优先）：用于归属项目、质保状态等字段
        grouped = {}
        for tl in traffic_lights:
            key = tl.intersection_id
            if key not in grouped:
                grouped[key] = tl
            elif ExcelExportService._should_keep_new_item(grouped[key], tl):
                grouped[key] = tl

        # 服役期限去重（选中状态优先）：仅用于设备服役时长计算
        grouped_service = {}
        for tl in traffic_lights:
            key = tl.intersection_id
            if key not in grouped_service:
                grouped_service[key] = tl
            elif ExcelExportService._should_keep_for_service(grouped_service[key], tl, tl_selected_map):
                grouped_service[key] = tl

        # 排序：基于通用去重结果，质保状态优先，再按归属项目
        status_order = {'过保': 0, '在保': 1}
        sorted_lights = sorted(grouped.values(), key=lambda tl: (
            status_order.get(ExcelExportService._get_warranty_status(tl.project_id), 2),
            ExcelExportService._get_project_info(tl.project_id)['name']
        ))
        
        row_idx = 1

        for tl in sorted_lights:
                project_info = ExcelExportService._get_project_info(tl.project_id)
                intersection_name = ExcelExportService._get_intersection_name(tl.intersection_id)
                intersection_type = ExcelExportService._get_intersection_type_name(tl.intersection_id)
                warranty_status = ExcelExportService._get_warranty_status(tl.project_id)
                
                # 设备服役时长：从服役期限去重结果计算
                usage_years = ''
                tl_service = grouped_service.get(tl.intersection_id)
                if tl_service:
                    service_project_info = ExcelExportService._get_project_info(tl_service.project_id)
                    if service_project_info['acceptance_date']:
                        acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                        usage_years = round((date.today() - acc_date).days / 365, 1)

                row = [
                    row_idx,
                    intersection_name,
                    intersection_type,
                    project_info['name'],
                    project_info['acceptance_date'],
                    project_info['warranty_period'],
                    project_info['warranty_expire_date'],
                    warranty_status,
                    project_info['builder'],
                    project_info['construction_unit'],
                    tl.signal_type or '',
                    tl.signal_count or 0,
                    tl.left_arrow_count or 0,
                    tl.straight_arrow_count or 0,
                    tl.right_arrow_count or 0,
                    tl.full_screen_count or 0,
                    tl.non_motor_count or 0,
                    tl.pedestrian_count or 0,
                    tl.countdown_timer_count or 0,
                    tl.radar_count or 0,
                    tl.guide_screen_count or 0,
                    tl.power_source or '',
                    usage_years
                ]
                ws.append(row)
                row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_electronic_police_sheet(wb):
        ws = wb.create_sheet('电子警察')
        headers = [
            '序号', '路口名称', '路口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍类型', '终端服务器数量', '正向抓拍数量', '反向抓拍数量',
            'LED灯', '爆闪灯', '监控球机数量', '信号检测器数量', '取网说明', '设备服役时长（年）'
        ]
        ws.append(headers)

        ep_list = db.session.query(ElectronicPolice).all()
        ep_intersections = db.session.query(Intersection).all()
        
        # 构建选中状态映射
        ep_selected_map = {}
        for inter in ep_intersections:
            if inter.selected_electronic_police_id:
                ep_selected_map[inter.selected_electronic_police_id] = True
        
        # 通用去重（质保日期优先）
        grouped = {}
        for ep in ep_list:
            key = ep.intersection_id
            if key not in grouped:
                grouped[key] = ep
            elif ExcelExportService._should_keep_new_item(grouped[key], ep):
                grouped[key] = ep

        # 服役期限去重（选中状态优先）：仅用于设备服役时长
        grouped_service = {}
        for ep in ep_list:
            key = ep.intersection_id
            if key not in grouped_service:
                grouped_service[key] = ep
            elif ExcelExportService._should_keep_for_service(grouped_service[key], ep, ep_selected_map):
                grouped_service[key] = ep

        # 排序：基于通用去重结果
        status_order = {'过保': 0, '在保': 1}
        sorted_eps = sorted(grouped.values(), key=lambda ep: (
            status_order.get(ExcelExportService._get_warranty_status(ep.project_id), 2),
            ExcelExportService._get_project_info(ep.project_id)['name']
        ))
        
        row_idx = 1

        for ep in sorted_eps:
                project_info = ExcelExportService._get_project_info(ep.project_id)
                intersection_name = ExcelExportService._get_intersection_name(ep.intersection_id)
                intersection_type = ExcelExportService._get_intersection_type_name(ep.intersection_id)
                warranty_status = ExcelExportService._get_warranty_status(ep.project_id)
                
                # 设备服役时长：从服役期限去重结果计算
                usage_years = ''
                ep_service = grouped_service.get(ep.intersection_id)
                if ep_service:
                    service_project_info = ExcelExportService._get_project_info(ep_service.project_id)
                    if service_project_info['acceptance_date']:
                        acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                        usage_years = round((date.today() - acc_date).days / 365, 1)

                row = [
                    row_idx,
                    intersection_name,
                    intersection_type,
                    project_info['name'],
                    project_info['acceptance_date'],
                    project_info['warranty_period'],
                    project_info['warranty_expire_date'],
                    warranty_status,
                    project_info['builder'],
                    project_info['construction_unit'],
                    ep.capture_type or '',
                    ep.terminal_server_count or 0,
                    ep.forward_capture_count or 0,
                    ep.reverse_capture_count or 0,
                    ep.led_light_count or 0,
                    ep.strobe_light_count or 0,
                    ep.ptz_count or 0,
                    ep.signal_detector_count or 0,
                    ep.network_source or '',
                    usage_years
                ]
                ws.append(row)
                row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_parking_enforcement_sheet(wb):
        ws = wb.create_sheet('违停球')
        headers = [
            '序号', '点位名称', '抓拍区域', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍机数量', '违停标牌数量', '监控标牌数量', '取电说明', '取网说明', '设备服役时长（年）'
        ]
        ws.append(headers)

        pe_list = db.session.query(ParkingEnforcement).all()
        pe_points = db.session.query(ParkingEnforcementPoint).all()
        
        # 构建选中状态映射
        pe_selected_map = {}
        for pt in pe_points:
            if pt.selected_project_id:
                pe_selected_map[pt.selected_project_id] = True
        
        grouped = {}
        for pe in pe_list:
            key = pe.point_id
            if key not in grouped:
                grouped[key] = pe
            elif ExcelExportService._should_keep_for_service(grouped[key], pe, pe_selected_map):
                grouped[key] = pe

        # 排序：质保状态优先（过保 > 在保 > 其他），再按归属项目
        status_order = {'过保': 0, '在保': 1}
        sorted_pes = sorted(grouped.values(), key=lambda pe: (
            status_order.get(ExcelExportService._get_warranty_status(pe.project_id), 2),
            ExcelExportService._get_project_info(pe.project_id)['name']
        ))
        
        row_idx = 1

        for pe in sorted_pes:
                project_info = ExcelExportService._get_project_info(pe.project_id)
                point_info = ExcelExportService._get_point_info(pe.point_id)
                warranty_status = ExcelExportService._get_warranty_status(pe.project_id)
                
                # 设备服役时长：从服役期限去重结果计算
                usage_years = ''
                pe_service = grouped_service.get(pe.point_id)
                if pe_service:
                    service_project_info = ExcelExportService._get_project_info(pe_service.project_id)
                    if service_project_info['acceptance_date']:
                        acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                        usage_years = round((date.today() - acc_date).days / 365, 1)

                row = [
                    row_idx,
                    point_info['name'],
                    point_info['area'],
                    project_info['name'],
                    project_info['acceptance_date'],
                    project_info['warranty_period'],
                    project_info['warranty_expire_date'],
                    warranty_status,
                    project_info['builder'],
                    project_info['construction_unit'],
                    pe.camera_count or 0,
                    pe.parking_sign_count or 0,
                    pe.monitor_sign_count or 0,
                    pe.power_source or '',
                    pe.network_source or '',
                    usage_years
                ]
                ws.append(row)
                row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_checkpoint_sheet(wb):
        ws = wb.create_sheet('卡口')
        headers = [
            '序号', '点位名称', '卡口类型', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '抓拍机数量', '爆闪灯数量', '测速雷达数量', '标牌数量', '取电说明', '取网说明', '设备服役时长（年）'
        ]
        ws.append(headers)

        cp_list = db.session.query(Checkpoint).all()
        cp_points = db.session.query(CheckpointPoint).all()
        
        # 构建选中状态映射
        cp_selected_map = {}
        for pt in cp_points:
            if pt.selected_project_id:
                cp_selected_map[pt.selected_project_id] = True
        
        # 通用去重（质保日期优先）
        grouped = {}
        for cp in cp_list:
            key = cp.point_id
            if key not in grouped:
                grouped[key] = cp
            elif ExcelExportService._should_keep_new_item(grouped[key], cp):
                grouped[key] = cp

        # 服役期限去重（选中状态优先）：仅用于设备服役时长
        grouped_service = {}
        for cp in cp_list:
            key = cp.point_id
            if key not in grouped_service:
                grouped_service[key] = cp
            elif ExcelExportService._should_keep_for_service(grouped_service[key], cp, cp_selected_map):
                grouped_service[key] = cp

        # 排序：基于通用去重结果
        status_order = {'过保': 0, '在保': 1}
        sorted_cps = sorted(grouped.values(), key=lambda cp: (
            status_order.get(ExcelExportService._get_warranty_status(cp.project_id), 2),
            ExcelExportService._get_project_info(cp.project_id)['name']
        ))
        
        row_idx = 1

        for cp in sorted_cps:
            project_info = ExcelExportService._get_project_info(cp.project_id)
            point_info = ExcelExportService._get_point_info(cp.point_id, 'checkpoint')
            warranty_status = ExcelExportService._get_warranty_status(cp.project_id)
            cp_type = point_info.get('area', '')
            
            # 设备服役时长：从服役期限去重结果计算
            usage_years = ''
            cp_service = grouped_service.get(cp.point_id)
            if cp_service:
                service_project_info = ExcelExportService._get_project_info(cp_service.project_id)
                if service_project_info['acceptance_date']:
                    acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                    usage_years = round((date.today() - acc_date).days / 365, 1)

            row = [
                row_idx,
                point_info['name'],
                cp_type,
                project_info['name'],
                project_info['acceptance_date'],
                project_info['warranty_period'],
                project_info['warranty_expire_date'],
                warranty_status,
                project_info['builder'],
                project_info['construction_unit'],
                cp.camera_count or 0,
                cp.strobe_light_count or 0,
                cp.radar_count or 0,
                cp.sign_count or 0,
                cp.power_source or '',
                cp.network_source or '',
                usage_years
            ]
            ws.append(row)
            row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_sky_net_sheet(wb):
        ws = wb.create_sheet('结构化相机')
        headers = [
            '序号', '点位名称', '监控区域', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位',
            '相机数量', '支架数量', '立杆数量', '挂箱数量', '补光灯数量', '音箱数量', '取电说明', '取网说明', '设备服役时长（年）'
        ]
        ws.append(headers)

        sn_list = db.session.query(SkyNet).all()
        sn_points = db.session.query(SkyNetPoint).all()
        
        # 构建选中状态映射
        sn_selected_map = {}
        for pt in sn_points:
            if pt.selected_project_id:
                sn_selected_map[pt.selected_project_id] = True
        
        # 通用去重（质保日期优先）
        grouped = {}
        for sn in sn_list:
            key = sn.point_id
            if key not in grouped:
                grouped[key] = sn
            elif ExcelExportService._should_keep_new_item(grouped[key], sn):
                grouped[key] = sn
        
        # 服役期限去重（选中状态优先）：仅用于设备服役时长
        grouped_service = {}
        for sn in sn_list:
            key = sn.point_id
            if key not in grouped_service:
                grouped_service[key] = sn
            elif ExcelExportService._should_keep_for_service(grouped_service[key], sn, sn_selected_map):
                grouped_service[key] = sn

        # 排序：基于通用去重结果
        status_order = {'过保': 0, '在保': 1}
        sorted_sns = sorted(grouped.values(), key=lambda sn: (
            status_order.get(ExcelExportService._get_warranty_status(sn.project_id), 2),
            ExcelExportService._get_project_info(sn.project_id)['name']
        ))
        
        row_idx = 1

        for sn in sorted_sns:
            project_info = ExcelExportService._get_project_info(sn.project_id)
            point = db.session.query(SkyNetPoint).get(sn.point_id)
            point_name = point.name if point else ''
            warranty_status = ExcelExportService._get_warranty_status(sn.project_id)
            
            # 设备服役时长：从服役期限去重结果计算
            usage_years = ''
            sn_service = grouped_service.get(sn.point_id)
            if sn_service:
                service_project_info = ExcelExportService._get_project_info(sn_service.project_id)
                if service_project_info['acceptance_date']:
                    acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                    usage_years = round((date.today() - acc_date).days / 365, 1)

            row = [
                row_idx,
                point_name,
                sn.camera_area or '',
                project_info['name'],
                project_info['acceptance_date'],
                project_info['warranty_period'],
                project_info['warranty_expire_date'],
                warranty_status,
                project_info['builder'],
                project_info['construction_unit'],
                sn.camera_count or 0,
                sn.bracket_count or 0,
                sn.pole_count or 0,
                sn.box_count or 0,
                sn.fill_light_count or 0,
                sn.speaker_count or 0,
                sn.power_source or '',
                sn.network_source or '',
                usage_years
            ]
            ws.append(row)
            row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _create_backend_device_sheet(wb):
        ws = wb.create_sheet('后端设备')
        headers = [
            '序号', '设备名称', '品牌型号', '设备类型', '设备数量', '归属项目', '项目验收日期', '项目质保期',
            '项目质保到期时间', '质保状态', '建设单位', '施工单位', '设备服役时长（年）'
        ]
        ws.append(headers)

        all_devices = db.session.query(BackendDevice).all()
        
        suffix_pattern = re.compile(r' \(\d+\)$')
        
        # 通用去重（质保日期优先）
        grouped = {}
        for d in all_devices:
            if suffix_pattern.search(d.name):
                continue
            key = d.name
            if key not in grouped:
                grouped[key] = d
            elif ExcelExportService._should_keep_new_item(grouped[key], d):
                grouped[key] = d
        
        # 服役期限去重（选中状态优先）：仅用于设备服役时长
        grouped_service = {}
        for d in all_devices:
            if suffix_pattern.search(d.name):
                continue
            key = d.name
            if key not in grouped_service:
                grouped_service[key] = d
            elif ExcelExportService._should_keep_for_service(grouped_service[key], d):
                grouped_service[key] = d

        # 排序：基于通用去重结果
        status_order = {'过保': 0, '在保': 1}
        sorted_devices = sorted(grouped.values(), key=lambda d: (
            status_order.get(ExcelExportService._get_warranty_status(d.project_id), 2),
            ExcelExportService._get_project_info(d.project_id)['name']
        ))
        
        row_idx = 1

        for d in sorted_devices:
            project_info = ExcelExportService._get_project_info(d.project_id)
            warranty_status = ExcelExportService._get_warranty_status(d.project_id)
            device_type = ExcelExportService._get_backend_device_type_name(d.id)
            
            # 设备服役时长：从服役期限去重结果计算
            usage_years = ''
            bd_service = grouped_service.get(d.name)
            if bd_service:
                service_project_info = ExcelExportService._get_project_info(bd_service.project_id)
                if service_project_info['acceptance_date']:
                    acc_date = date.fromisoformat(service_project_info['acceptance_date'])
                    usage_years = round((date.today() - acc_date).days / 365, 1)

            row = [
                row_idx,
                d.name or '',
                d.model or '',
                device_type,
                d.quantity or 1,
                project_info['name'],
                project_info['acceptance_date'],
                project_info['warranty_period'],
                project_info['warranty_expire_date'],
                warranty_status,
                project_info['builder'],
                project_info['construction_unit'],
                usage_years
            ]
            ws.append(row)
            row_idx += 1

        ExcelExportService._auto_adjust_column_width(ws)

    @staticmethod
    def _auto_adjust_column_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
